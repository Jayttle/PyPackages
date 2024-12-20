import os
import shutil
import threading
import concurrent.futures
import subprocess
import gzip
from enum import Enum
from collections import defaultdict, OrderedDict
from openpyxl import Workbook
from openpyxl.styles import Alignment
from datetime import datetime, timedelta
from pathlib import Path
from time import sleep
from typing import Union, Optional
from JayttleProcess import CommonDecorator


class RinexFileType(Enum):
    O = 1
    N = 2
    Unkonw  = 3


class FileFormat(Enum):
    CRX = 1
    RNX = 2
    Unkonw  = 3


class StationType(Enum):
    RoverStation = 1
    BaseStation = 2
    Unkonw = 3


class RinexFileInfo:
    def __init__(self, input_path: str, ftp_file: bool):
        self.ftp_file: bool = ftp_file  # 是否来自FTP
        self.input_path: str = input_path  # 输入路径
        self.file_info: Path = Path(input_path)  # 文件信息对象
        self.file_name: str = self.file_info.name  # 文件名
        self.station_name: str = self.file_name[:3]  # 站点名称，前三个字符
        self.marker_name: str = self.file_name[:4]  # 标记名称，前四个字符
        self.station_id: int = int(self.station_name[2])  # 站点ID，从站点名称中第三个字符转换而来
        self.station_type: Union[StationType, str] = {  # 站点类型，根据站点名称的第一个字符确定，如果未知则为 Unknown
            'R': StationType.RoverStation,
            'B': StationType.BaseStation
        }.get(self.station_name[0], StationType.Unkonw)
        self.receiver_id: int = int(self.file_name[3])  # 接收器ID，从文件名第四个字符转换而来
        split = self.file_name.split('_')
        self.start_gps_time_str: str = split[2]  # 起始GPS时间字符串，从文件名中获取
        self.start_gps_time: datetime = self.get_time_from_string(self.start_gps_time_str)  # 起始GPS时间，使用自定义函数从字符串中获取
        self.duration_str: str = split[3]  # 持续时间字符串，从文件名中获取
        self.duration: timedelta = self.get_duration_from_string(self.duration_str)  # 持续时间，使用自定义函数从字符串中获取
        self.time_str: str = f"{self.start_gps_time_str}_{self.duration_str}"  # 时间字符串，包含起始GPS时间和持续时间
        file_type = split[-1][1]
        self.file_type: Union[RinexFileType, str] = {  # 文件类型，根据文件名中的第二个字符确定，如果未知则为 Unknown
            'O': RinexFileType.O,
            'N': RinexFileType.N
        }.get(file_type, RinexFileType.Unkonw)
        split1 = self.file_name.split('.')
        self.info_str: str = split1[0]  # 信息字符串，从文件名中获取
        compressed = split1[-1].lower()
        self.compressed: bool = compressed == "zip" or compressed == "z"  # 是否为压缩文件
        self.format: Union[FileFormat, str] = {    # 文件格式，根据文件扩展名确定，如果未知则为 Unknown
            "crx": FileFormat.CRX,
            "rnx": FileFormat.RNX
        }.get(split1[1].lower(), FileFormat.Unkonw)


    @staticmethod
    def get_time_from_string(time_str: str) -> datetime:
        """
        将字符串形式的时间转换为 datetime 对象。

        参数：
            time_str (str): 表示时间的字符串，格式为 'YYYYDDDHHMM'，其中：
                YYYY: 年份
                DDD: 年份中的第几天
                HH: 小时
                MM: 分钟

        返回：
            datetime: 表示转换后的时间的 datetime 对象。
        """
        year = int(time_str[:4])  # 年份
        day_of_year = int(time_str[4:7])  # 年份中的第几天
        hour = int(time_str[7:9])  # 小时
        minute = int(time_str[9:11])  # 分钟
        return datetime(year, 1, 1) + timedelta(days=day_of_year - 1, hours=hour, minutes=minute)


    @staticmethod
    def get_duration_from_string(duration_str: str) -> timedelta:
        """
        将字符串形式的时长转换为 timedelta 对象。

        参数：
            duration_str (str): 表示时长的字符串，格式为 'X[D/H/M/S]'，其中：
                X: 数字，表示时长的数量
                D: 天
                H: 小时
                M: 分钟
                S: 秒

        返回：
            timedelta: 表示转换后的时长的 timedelta 对象。
        """
        days = int(duration_str[:-1])
        unit = duration_str[-1]
        if unit == 'D':
            return timedelta(days=days)
        elif unit == 'H':
            return timedelta(hours=days)
        elif unit == 'M':
            return timedelta(minutes=days)
        elif unit == 'S':
            return timedelta(seconds=days)

    @staticmethod
    def get_string_from_duration(duration: timedelta) -> str:
        """
        将 timedelta 对象转换为字符串形式的时长。

        参数：
            duration (timedelta): 表示时长的 timedelta 对象。

        返回：
            str: 表示转换后的字符串形式的时长，格式为 '[X]D/H/M/S'，其中：
                X: 数字，表示时长的数量
                D: 天
                H: 小时
                M: 分钟
                S: 秒
        """
        if duration < timedelta(minutes=1):
            return f"{duration.total_seconds():.0f}S"
        elif duration < timedelta(hours=1):
            return f"{duration.total_seconds() / 60:.0f}M"
        elif duration < timedelta(days=1):
            return f"{duration.total_seconds() / 3600:.0f}H"
        else:
            return f"{duration.days}D"


def get_rnx_files(directory_path: str) -> list[RinexFileInfo]:
    """
    获取指定目录下的所有 Rinex 文件的信息。

    参数：
        directory_path (str): Rinex 文件所在目录的路径。

    返回：
        list[RinexFileInfo]: Rinex 文件信息的列表。
    """
    rnx_files = []
    for file_name in os.listdir(directory_path):
        if file_name.endswith(".rnx"):
            file_path = os.path.join(directory_path, file_name)
            rnx_file = RinexFileInfo(file_path, ftp_file=False)
            rnx_files.append(rnx_file)
    return rnx_files


class MergeFiles:
    @staticmethod
    @CommonDecorator.log_function_call
    def merge_files(files: list[RinexFileInfo], merge_path: str, rewrite: bool = False) -> bool:
        """
        合并多个 Rinex 文件。

        参数：
            files (list[RinexFileInfo]): 要合并的 Rinex 文件信息的列表。
            merge_path (str): 合并后文件存储的路径。
            rewrite (bool, 可选): 如果合并后的文件已存在，是否覆盖。默认为 False。

        返回：
            bool: 表示合并是否成功的布尔值。

        注意：
            如果文件列表为空或只有一个文件，则无法进行合并，会返回 False。
        """
        if files is None or len(files) <= 1:
            return False
        if not os.path.exists(merge_path):
            os.makedirs(merge_path)

        first = files[0]
        last = files[-1]

        split = first.file_name.split('_')
        start_time = first.start_gps_time
        start_time_str = first.start_gps_time_str
        duration = (last.start_gps_time - first.start_gps_time) + last.duration
        duration_str = RinexFileInfo.get_string_from_duration(duration)
        split[2] = start_time_str
        split[3] = duration_str
        merge_name = '_'.join(split)

        merge_path = os.path.join(merge_path, start_time_str)
        if not os.path.exists(merge_path):
            os.makedirs(merge_path)

        merge_full_name = os.path.join(merge_path, merge_name)  # Modify this line
        if os.path.exists(merge_full_name):
            if not rewrite:
                return True
            else:
                os.remove(merge_full_name)

        # 构建正确的参数列表
        file_names = [f.file_info for f in files]
        command = [r"D:\Program Files (x86)\Software\OneDrive\C#\WPF\Cableway.Download\Rinex\gfzrnx_2.1.0_win64.exe"]
        options = ["-finp"] + file_names + ["-kv", "-fout", merge_full_name, "-vo", "3.02"]

        # 执行外部程序
        process = subprocess.Popen(
            command + options,
            shell=False,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE
        )
        stdout, stderr = process.communicate()

        sleep(0.1)
        return os.path.exists(merge_full_name)


def get_time_from_str_version2(time_str: str) -> datetime:
    """
    从时间字符串中提取 datetime 对象。

    参数：
        time_str (str): 表示时间的字符串，格式为 "YYYYMMDDHH"。

    返回：
        datetime: 时间对应的 datetime 对象。
    """
    year = int(time_str[:4])  # 年份
    month = int(time_str[4:6])  # 月份
    day = int(time_str[6:8])  # 天
    hour = int(time_str[8:10])  # 小时
    return datetime(year, month, day, hour) 


def get_special_datetimes(directory: str) -> list:
    """
    获取指定目录下特定时间范围内的 datetime 列表。

    参数：
        directory (str): 要搜索的目录路径。

    返回：
        list: 指定时间范围内的 datetime 对象列表。
    """
    special_datetimes = []
    for folder_name in os.listdir(directory):
        try:
            dt = get_time_from_str_version2(folder_name)
            # 检查是否在指定时间范围内
            if dt.time() >= datetime.strptime("12:00:00", "%H:%M:%S").time() and dt.time() <= datetime.strptime("23:00:00", "%H:%M:%S").time():
                special_datetimes.append(dt)
        except ValueError:
            pass  # 文件夹名称格式不正确，跳过
    return special_datetimes


def filter_and_group_folders(directory: str) -> dict:
    """
    根据文件夹名称对文件夹进行过滤和分组。

    参数：
        directory (str): 要处理的文件夹路径。

    返回：
        dict: 包含分组信息的字典，键为日期的前八位数字，值为包含满足条件的小时数的列表。
    """
    folder_groups = {}

    for folder_name in os.listdir(directory):
        if folder_name[-2:].isdigit():
            last_two_digits = int(folder_name[-2:])
            first_eight_digits = int(folder_name[:8])

            if (last_two_digits >= 12 and last_two_digits <= 23):
                if first_eight_digits not in folder_groups:
                    folder_groups[first_eight_digits] = []
                folder_groups[first_eight_digits].append(last_two_digits)

    # 移除不满足条件的分组
    folder_groups = {key: value for key, value in folder_groups.items() if len(value) >= 4}

    return folder_groups


def group_by_hour(file_list: list[RinexFileInfo]) -> dict[str, list[RinexFileInfo]]:
    """
    将 RinexFileInfo 对象列表按照小时范围进行分组。

    参数：
        file_list (list[RinexFileInfo]): 包含 RinexFileInfo 对象的列表。

    返回：
        dict[str, list[RinexFileInfo]]: 按小时范围分组后的字典，键为小时范围字符串，值为对应的 RinexFileInfo 对象列表。
    """
    groups = {"12-15": [], "16-19": [], "20-23": []}
    for file_info in file_list:
        hour_key = int(file_info.start_gps_time.strftime("%H"))
        if 12 <= hour_key < 16:
            groups["12-15"].append(file_info)
        elif 16 <= hour_key < 20:
            groups["16-19"].append(file_info)
        elif 20 <= hour_key <= 23:
            groups["20-23"].append(file_info)
    return groups


def merge_files_by_hour(groups: dict[str, list[RinexFileInfo]], merge_path: str) -> None:
    """
    将按小时分组的 RinexFileInfo 对象列表合并到指定路径。

    参数：
        groups (dict[str, list[RinexFileInfo]]): 按小时范围分组的 RinexFileInfo 对象字典。
        merge_path (str): 合并后的文件存储路径。
    """
    for hour_range, file_group in groups.items():
        MergeFiles.merge_files(file_group, merge_path)


def merge_files_by_hour_multithread(groups: dict[str, list[RinexFileInfo]], merge_path: str) -> None:
    """
    使用多线程将按小时分组的 RinexFileInfo 对象列表合并到指定路径。

    参数：
        groups (dict[str, list[RinexFileInfo]]): 按小时范围分组的 RinexFileInfo 对象字典。
        merge_path (str): 合并后的文件存储路径。
    """
    threads = [threading.Thread(target=MergeFiles.merge_files, args=(file_group, merge_path)) for file_group in groups.values()]
    for thread in threads:
        thread.start()
    for thread in threads:
        thread.join()


def merge_files_by_hour_threadpool(groups: dict[str, list[RinexFileInfo]], merge_path: str) -> None:
    """
    使用线程池将按小时分组的 RinexFileInfo 对象列表合并到指定路径。

    参数：
        groups (dict[str, list[RinexFileInfo]]): 按小时范围分组的 RinexFileInfo 对象字典。
        merge_path (str): 合并后的文件存储路径。
    """
    # 使用 ThreadPoolExecutor 来管理线程池
    with concurrent.futures.ThreadPoolExecutor(max_workers=6) as executor:
        # 创建一个将 future 对象映射到小时范围的字典
        futures = {executor.submit(MergeFiles.merge_files, file_group, merge_path): hour_range for hour_range, file_group in groups.items()}
        # 遍历 future 对象，直到它们完成
        for future in concurrent.futures.as_completed(futures):
            try:
                # 尝试获取 future 的结果，如果出现异常则捕获并打印
                future.result()
            except Exception as e:
                print(f"发生异常：{e}")


def process_directory(directory_path: str, merge_path: str):
    """
    处理指定目录中的文件。

    参数：
    - directory_path：要处理的目录路径。
    - merge_path：合并文件后的保存路径。

    返回：
    无。
    """
    rnx_files: list[RinexFileInfo] = get_rnx_files(directory_path)
    o_files: list[RinexFileInfo] = [rnx_file for rnx_file in rnx_files if rnx_file.file_type == RinexFileType.O]
    n_files: list[RinexFileInfo] = [rnx_file for rnx_file in rnx_files if rnx_file.file_type == RinexFileType.N]
    
    # 按小时分组
    grouped_o_files: dict[str, list[RinexFileInfo]] = group_by_hour(o_files)
    grouped_n_files: dict[str, list[RinexFileInfo]] = group_by_hour(n_files)

    merge_files_by_hour_multithread(grouped_o_files, merge_path)
    merge_files_by_hour_multithread(grouped_n_files, merge_path)


def delete_directories(directories: list[str]):
    """
    删除指定的文件夹。

    参数：
    - directories：包含要删除的文件夹路径的列表。

    返回：
    无。
    """
    for directory in directories:
        if os.path.exists(directory):
            shutil.rmtree(directory)
            print(f"Deleted directory: {directory}")
        else:
            print(f"Directory does not exist: {directory}")


def proj_merge_rnx():
    directory_paths = [
        r"D:\Ropeway\GNSS\FTP\B011",
        r"D:\Ropeway\GNSS\FTP\B021",
        r"D:\Ropeway\GNSS\FTP\R031"
    ]
    merge_path = r"D:\Ropeway\GNSS\FTP\Merge"
    for directory_path in directory_paths:
        process_directory(directory_path, merge_path)

    # delete_directories(directory_paths)
    
# proj_merge_rnx()


def read_rinex_file_info(file_path: str) -> RinexFileInfo:
    """
    从 Rinex 文件名中解析 RinexFileInfo 对象。

    参数：
        file_path (str): Rinex 文件的路径。

    返回：
        RinexFileInfo: 解析得到的 Rinex 文件信息。
    """
    file_name = os.path.basename(file_path)
    file_name = file_name.replace(".crx.Z", ".rnx").replace(".rnx.Z", ".rnx")  # 统一替换文件类型后缀
    ftp_file = False  # 假设文件不是来自 FTP
    return RinexFileInfo(file_path, ftp_file)


def read_rinex_files_info(file_list_path: str) -> list[RinexFileInfo]:
    """
    从文件列表中逐行读取 Rinex 文件信息并返回 RinexFileInfo 对象列表。

    参数：
        file_list_path (str): 包含 Rinex 文件名的文本文件路径。

    返回：
        list[RinexFileInfo]: Rinex 文件信息的列表。
    """
    rnx_files_info = []
    with open(file_list_path, 'r') as file:
        for line in file:
            line = line.strip()  # 去除行末的换行符
            line = line.replace(".crx.Z", ".rnx").replace(".rnx.Z", ".rnx")  # 统一替换文件类型后缀
            if line.endswith(".rnx"):  # 确保是 Rinex 文件
                rnx_file_info = read_rinex_file_info(line)
                rnx_files_info.append(rnx_file_info)
    return rnx_files_info


def read_rinex_files_info_version2(file_list_path: str) -> list[RinexFileInfo]:
    """
    从文件列表中逐行读取 Rinex 文件信息并返回 RinexFileInfo 对象列表。

    参数：
        file_list_path (str): 包含 Rinex 文件名的文本文件路径。

    返回：
        list[RinexFileInfo]: Rinex 文件信息的列表。
    """
    rnx_files_info = []
    with open(file_list_path, 'r') as file:
        for line in file:
            line = line.strip()  # 去除行末的换行符
            line = line.replace(".crx.Z", ".rnx").replace(".rnx.Z", ".rnx")  # 统一替换文件类型后缀
            if line.endswith(".rnx") or line.endswith(".crx.Z") or line.endswith(".rnx.Z"):  # 确保是 Rinex 文件
                rnx_file_info = read_rinex_file_info(line)
                rnx_files_info.append(rnx_file_info)
    return rnx_files_info


def get_rnx_files_dict(directory_path: str) -> dict[datetime, list[RinexFileInfo]]:
    """
    获取指定目录下的所有 Rinex 文件的信息，并以起始GPS时间为键，RinexFileInfo 对象为值的列表形式返回。

    参数：
        directory_path (str): Rinex 文件所在目录的路径。

    返回：
        Dict[datetime, List[RinexFileInfo]]: Rinex 文件信息的字典，以起始GPS时间为键，对应的 RinexFileInfo 对象列表为值。
    """
    rinex_files_dict = {}
    for file_name in os.listdir(directory_path):
        if file_name.endswith(".rnx") or file_name.endswith(".crx"):
            file_path = os.path.join(directory_path, file_name)
            rinex_file = RinexFileInfo(file_path, ftp_file=False)
            start_gps_time = rinex_file.start_gps_time
            if start_gps_time not in rinex_files_dict:
                rinex_files_dict[start_gps_time] = []
            rinex_files_dict[start_gps_time].append(rinex_file)
    return rinex_files_dict


def get_rnx_files_dict_date(directory_path: str) -> dict[datetime.date, list[RinexFileInfo]]:
    """
    获取指定目录下的所有 Rinex 文件的信息，并以日期为键，RinexFileInfo 对象为值的列表形式返回。

    参数：
        directory_path (str): Rinex 文件所在目录的路径。

    返回：
        Dict[date, List[RinexFileInfo]]: Rinex 文件信息的字典，以日期为键，对应的 RinexFileInfo 对象列表为值。
    """
    rinex_files_dict = {}
    for file_name in os.listdir(directory_path):
        if file_name.endswith(".rnx") or file_name.endswith(".crx"):
            file_path = os.path.join(directory_path, file_name)
            rinex_file = RinexFileInfo(file_path, ftp_file=False)
            start_gps_time = rinex_file.start_gps_time
            start_date = start_gps_time.date()
            if start_date not in rinex_files_dict:
                rinex_files_dict[start_date] = []
            rinex_files_dict[start_date].append(rinex_file)
    return rinex_files_dict


def create_marker_name_excel(marker_name_files_per_day_1h: dict[datetime.date, dict], marker_name: str, file_path: str) -> None:
    # 创建一个新的工作簿
    wb = Workbook()

    # 写入每天中持续时间为 01 小时的文件数量，包括标题行
    ws_1h = wb.active
    ws_1h.title = "Files per Day (01H)"
    ws_1h['A1'] = "Date"
    ws_1h['B1'] = "File Count"
    ws_1h['C1'] = "Percentage"

    # 设置标题行样式
    for cell in ['A1', 'B1', 'C1']:
        ws_1h[cell].alignment = Alignment(horizontal='center')

    # 获取当前 marker_name 的最大文件数量
    max_count = max(marker_name_files_per_day_1h[marker_name].values())

    # 遍历对应 marker_name 的数据，写入到工作表中
    for idx, (date, count) in enumerate(marker_name_files_per_day_1h[marker_name].items(), start=2):
        ws_1h[f'A{idx}'] = date.strftime('%Y-%m-%d')
        ws_1h[f'B{idx}'] = count
        # 计算百分比
        percentage = count / max_count * 100 if max_count != 0 else 0
        ws_1h[f'C{idx}'] = f"{percentage:.2f}%"

    # 保存工作簿
    wb.save(file_path)


def Proj_export_excel(rinex_files_info: list[RinexFileInfo]):
    # 创建一个嵌套的 defaultdict 来存储不同 marker_name 的每天中持续时间为 01 小时的文件数量
    marker_name_files_per_day_1h = defaultdict(lambda: defaultdict(int))
    # 遍历 RinexFileInfo 列表
    for file_info in rinex_files_info:
        # 获取 marker_name
        marker_name = file_info.marker_name
        
        # 获取持续时间字符串
        duration_str = file_info.duration_str

        # 检查持续时间是否为 01H
        if duration_str[-3:] == "01H":
            # 获取文件起始时间的日期
            start_date: datetime.date = file_info.start_gps_time.date()

            # 增加对应 marker_name 和日期的文件数量
            marker_name_files_per_day_1h[marker_name][start_date] += 1

    # 输出到 Excel 表格中
    for marker_name in marker_name_files_per_day_1h.keys():
        file_path = f"{marker_name}_statistics.xlsx"  # 根据 marker_name 构造文件名
        create_marker_name_excel(marker_name_files_per_day_1h, marker_name, file_path)


def create_excel_file(files_per_day_1h: dict[str, int], duration_str_counts: dict[str, int]) -> None:
    # 创建一个新的工作簿
    wb = Workbook()

    # 写入每天中持续时间为 01 小时的文件数量，包括标题行
    ws_1h = wb.create_sheet(title="Files per Day (01H)")
    ws_1h['A1'] = "Date"
    ws_1h['B1'] = "File Count"

    # 设置标题行样式
    for cell in ['A1', 'B1']:
        ws_1h[cell].alignment = Alignment(horizontal='center')

    for idx, (date, count) in enumerate(files_per_day_1h.items(), start=2):
        ws_1h[f'A{idx}'] = date.strftime('%Y-%m-%d')
        ws_1h[f'B{idx}'] = count

    # 设置列宽和单元格对齐方式
    for col in ['A', 'B']:
        ws_1h.column_dimensions[col].width = 13.88
        for row in ws_1h.iter_rows(min_row=2, max_row=len(files_per_day_1h)+1, min_col=ws_1h[col][0].column, max_col=ws_1h[col][0].column):
            for cell in row:
                cell.alignment = Alignment(horizontal='center')

    # 写入持续时间字符串的种类及对应的文件数量，包括标题行
    ws_duration = wb.create_sheet(title="Duration Counts")
    ws_duration['A1'] = "Duration Type"
    ws_duration['B1'] = "File Count"

    # 设置标题行样式
    for cell in ['A1', 'B1']:
        ws_duration[cell].alignment = Alignment(horizontal='center')

    for idx, (duration_str, count) in enumerate(duration_str_counts.items(), start=2):
        ws_duration[f'A{idx}'] = duration_str
        ws_duration[f'B{idx}'] = count

    # 设置列宽和单元格对齐方式
    for col in ['A', 'B']:
        ws_duration.column_dimensions[col].width = 13.88
        for row in ws_duration.iter_rows(min_row=2, max_row=len(duration_str_counts)+1, min_col=ws_duration[col][0].column, max_col=ws_duration[col][0].column):
            for cell in row:
                cell.alignment = Alignment(horizontal='center')

    # 删除默认的工作表
    default_sheet = wb['Sheet']
    wb.remove(default_sheet)

    # 保存工作簿
    excel_file_path = "statistics.xlsx"
    wb.save(excel_file_path)


def count_files_per_day(rinex_files_info: list[RinexFileInfo]) -> dict[datetime.date, dict[str, int]]:
    """
    统计每天文件数，并按日期和标记站名返回文件数的字典。

    参数：
        rinex_files_info (list[RinexFileInfo]): Rinex 文件信息的列表。

    返回：
        OrderedDict[date, OrderedDict[str, int]]: 每天文件数的有序字典，键为日期，值为另一个有序字典，
                                                   其中键为标记站名，值为文件数。
    """
    # 初始化一个包含所有标记站名的列表
    all_marker_names = ['R031', 'R032', 'R051', 'R052', 'R071', 'R072', 'R081', 'R082', 'B011', 'B021']

    files_per_day = defaultdict(lambda: defaultdict(int))
    
    for rinex_file in rinex_files_info:
        file_date = rinex_file.start_gps_time.date()  # 获取文件的日期信息
        marker_name = rinex_file.marker_name  # 获取标记站名
        
        # 获取持续时间字符串
        duration_str = rinex_file.duration_str

        # 检查持续时间是否为 01H
        if duration_str[-3:] == "01H":
            # 在字典中增加或更新日期和标记站名对应的文件数
            files_per_day[file_date][marker_name] += 1
    
    # 确保每个日期都包含所有的标记站名，将缺失的文件数填充为0
    for files in files_per_day.values():
        for marker_name in all_marker_names:
            files[marker_name] = files.get(marker_name, 0)
    
    # 对结果进行排序
    sorted_files_per_day = OrderedDict(sorted(files_per_day.items()))
    for files in sorted_files_per_day.values():
        sorted_files = OrderedDict(sorted(files.items()))
        files.clear()
        files.update(sorted_files)
    
    return sorted_files_per_day


def count_files_in_hour_range(rinex_files_info: list[RinexFileInfo], start_hour: int, end_hour: int) -> dict[datetime.date, dict[str, int]]:
    """
    统计每天指定时间范围内文件的数量，并按日期和标记站名返回文件数的字典。

    参数：
        rinex_files_info (list[RinexFileInfo]): Rinex 文件信息的列表。
        start_hour (int): 起始小时。
        end_hour (int): 结束小时。

    返回：
        OrderedDict[date, OrderedDict[str, int]]: 每天文件数的有序字典，键为日期，值为另一个有序字典，
                                                   其中键为标记站名，值为文件数。
    """
    # 初始化一个包含所有标记站名的列表
    all_marker_names = ['R031', 'R032', 'R051', 'R052', 'R071', 'R072', 'R081', 'R082', 'B011', 'B021']

    files_per_day = defaultdict(lambda: defaultdict(int))
    
    for rinex_file in rinex_files_info:
        file_date = rinex_file.start_gps_time.date()  # 获取文件的日期信息
        file_hour = rinex_file.start_gps_time.hour  # 获取文件的小时信息
        marker_name = rinex_file.marker_name  # 获取标记站名
        

        # 获取持续时间字符串
        duration_str = rinex_file.duration_str

        # 检查持续时间是否为 01H
        if duration_str[-3:] == "01H":
            # 检查文件的小时是否在指定范围内
            if start_hour <= file_hour <= end_hour:
                # 在字典中增加或更新日期和标记站名对应的文件数
                files_per_day[file_date][marker_name] += 1
    
    # 确保每个日期都包含所有的标记站名，将缺失的文件数填充为0
    for files in files_per_day.values():
        for marker_name in all_marker_names:
            files[marker_name] = files.get(marker_name, 0)
    
    # 对结果进行排序
    sorted_files_per_day = OrderedDict(sorted(files_per_day.items()))
    for files in sorted_files_per_day.values():
        sorted_files = OrderedDict(sorted(files.items()))
        files.clear()
        files.update(sorted_files)
    
    return sorted_files_per_day


def find_dates_with_specific_file_count(files_in_hour_range: dict[datetime.date, dict[str, int]], marker_names: list[str], count: int) -> list[datetime.date]:
    """
    查找指定标记站名文件数同时为特定数量的日期列表。

    参数：
        files_in_hour_range (dict[datetime.date, dict[str, int]]): 包含文件数信息的字典。
        marker_names (list[str]): 要检查的标记站名列表。
        count (int): 要检查的文件数。

    返回：
        list[datetime.date]: 符合条件的日期列表。
    """
    dates_with_specific_count = []

    for date, file_counts in files_in_hour_range.items():
        # 检查每个标记站名的文件数是否满足要求
        if all(file_counts.get(marker_name, 0) == count for marker_name in marker_names):
            dates_with_specific_count.append(date)

    return dates_with_specific_count


def proj_1(): 
    # 使用示例：
    directory_path = "D:\\Ropeway\\GNSS\\"
    file_list_path = os.path.join(directory_path, "all_files.txt")
    # file_list_path = os.path.join(directory_path, "towerbase2_files.txt")
    rinex_files_info = read_rinex_files_info(file_list_path)

    # 使用示例：
    start_hour = 16
    end_hour = 19
    files_in_hour_range = count_files_in_hour_range(rinex_files_info, start_hour, end_hour)

    specified_marker_names = ['R031', 'B011', 'B021']
    # 要检查的文件数
    specified_count = 8


    print(f"{start_hour}~{end_hour}可用天数:")
    # 覦标记站名列表
    for item in ["R031", "R032", "R051", "R052", "R071", "R072", "R081", "R082"]:
        specified_marker_names[0] = item
        # 查找指定标记站名文件数同时为8的日期列表
        dates_with_specific_count = find_dates_with_specific_file_count(files_in_hour_range, specified_marker_names, specified_count)
        print(f"{item}:{len(dates_with_specific_count)}")


def creat_excel(export_data: dict[datetime.date, dict[str, int]]) -> None:
    # 创建一个新的Excel工作簿
    wb = Workbook()
    ws = wb.active

    # 写入表头
    header = ["日期", "B011", "B021", "R031", "R032", "R051", "R052", "R071", "R072", "R081", "R082"]
    ws.append(header)

    # 将数据写入表格
    for date, file_counts in export_data.items():
        row = [date.strftime("%Y-%m-%d")] + [file_counts.get(marker, 0) for marker in header[1:]]
        ws.append(row)

    # 保存Excel文件
    output_excel_path = "D:\\Ropeway\\GNSS\\files_in_hour_range.xlsx"
    wb.save(output_excel_path)


def write_rinex_files_to_txt(rinex_files_info: list[RinexFileInfo], 
                              marker_name: str, 
                              dates: list[datetime.date], 
                              output_file_path: str,
                              start_hour: int,
                              end_hour: int) -> None:
    with open(output_file_path, 'w') as output_file:
        for rinex_file_info in rinex_files_info:
            file_hour = rinex_file_info.start_gps_time.hour
            file_date = rinex_file_info.start_gps_time.date()
            if rinex_file_info.marker_name == marker_name and file_date in dates and start_hour <= file_hour <= end_hour:
                # 替换文件名后缀
                file_name = rinex_file_info.file_name.replace("MO.rnx", "MO.crx.Z").replace("MN.rnx", "MN.rnx.Z")
                output_file.write(f"{file_name}\n")


def crx_to_rnx(file_info: RinexFileInfo, rnx_path: Optional[str] = None, delete_crx_file: bool = True, overwrite_repeated: bool = False) -> Optional[RinexFileInfo]:
    if(file_info.format == FileFormat.CRX):
        process = subprocess.Popen(["D:\Program Files (x86)\Software\OneDrive\C#\windows_C#\Cableway.Net7\Cableway.Download\bin\Release\net8.0-windows\Rinex\CRX2RNX.exe",
                                    file_info.file_path],
                                    stdout=subprocess.PIPE,
                                    stderr=subprocess.PIPE)
        process.wait()
        # 检查进程是否成功执行
        if process.returncode == 0:
            if delete_crx_file and file_info.exists:
                os.remove(file_info.file_path)
        else:
            print("CRX to RNX conversion failed with error:", process.stderr.read())   


def gzip_folder_path(folder_path: str = r"D:\Ropeway\GNSS\R031") -> None:
    # 指定文件夹路径
    # 遍历文件夹中的所有文件
    for file_name in os.listdir(folder_path):
        # 检查文件是否以 .Z 结尾
        if file_name.endswith(".Z"):
            # 构建文件的完整路径
            file_path = os.path.join(folder_path, file_name)
            # 解压缩文件
            with gzip.open(file_path, 'rb') as f_in:
                # 读取解压后的内容
                uncompressed_content = f_in.read()
            
            # 构建解压后文件的文件名（去除 .Z 后缀）
            uncompressed_file_name = file_name[:-2]
            # 构建解压后文件的完整路径
            uncompressed_file_path = os.path.join(folder_path, uncompressed_file_name)
            
            # 将解压后的内容写入新文件
            with open(uncompressed_file_path, 'wb') as f_out:
                f_out.write(uncompressed_content)
            
            # 删除原始的 .Z 文件
            os.remove(file_path)


def find_8files():
    # 使用示例：
    directory_path = "D:\\Ropeway\\GNSS\\"
    file_list_path = os.path.join(directory_path, "all_files.txt")
    # file_list_path = os.path.join(directory_path, "towerbase2_files.txt")
    rinex_files_info = read_rinex_files_info(file_list_path)

    # 使用示例：
    start_hour = 16
    end_hour = 19
    files_in_hour_range = count_files_in_hour_range(rinex_files_info, start_hour, end_hour)

    specified_marker_names = ['R031', 'B011', 'B021']
    # 要检查的文件数
    specified_count = 8


    print(f"{start_hour}~{end_hour}可用天数:")

    # 查找指定标记站名文件数同时为8的日期列表
    dates_with_specific_count = find_dates_with_specific_file_count(files_in_hour_range, specified_marker_names, specified_count)
    item = 'B021'
    print(f"{item}:{len(dates_with_specific_count)}")
    # 示例用法
    output_file_path = f"D:\\Ropeway\\GNSS\\R031_output_{item}.txt"
    marker_name = item
    write_rinex_files_to_txt(rinex_files_info, marker_name, dates_with_specific_count, output_file_path, start_hour, end_hour)


def crx_to_rnx(rnx_files: dict[datetime, list[RinexFileInfo]], delete_crx_file: bool = True, overwrite: bool = True) -> None:
    for start_time, rinex_list in rnx_files.items():
        if len(rinex_list) == 3:
            # 删除 file_type 为 RinexFileType.O，format 为 FileFormat.RNX 的文件
            for rnx_info in rinex_list:
                if rnx_info.file_type == RinexFileType.O and rnx_info.format == FileFormat.RNX:
                    if os.path.exists(rnx_info.file_info):
                        os.remove(rnx_info.file_info)

        # 对剩余的 format 为 FileFormat.CRX 的文件进行转换
        for rnx_info in rinex_list:
            if rnx_info.format == FileFormat.CRX:
                process = subprocess.Popen([r"C:\Program Files\Trimble\Trimble Business Center\CRX2RNX.exe",
                                            rnx_info.file_info],
                                            stdout=subprocess.PIPE,
                                            stderr=subprocess.PIPE)
                process.wait()
                # 检查进程是否成功执行
                if process.returncode == 0:
                    # 如果允许删除原始 CRX 文件，则删除它
                    if delete_crx_file and os.path.exists(rnx_info.file_info):
                        os.remove(rnx_info.file_info)
                else:
                    print("CRX to RNX conversion failed with error:", process.stderr.read())




def merge_files_threadpool_version2(file_groups: list[list[RinexFileInfo]], merge_path: str) -> None:
    """
    使用线程池将 RinexFileInfo 对象列表合并到指定路径。

    参数：
        file_groups (list[list[RinexFileInfo]]): RinexFileInfo 对象列表的列表。
        merge_path (str): 合并后的文件存储路径。
    """
    # 使用 ThreadPoolExecutor 来管理线程池
    with concurrent.futures.ThreadPoolExecutor(max_workers=12) as executor:
        # 创建一个将 future 对象映射到 RinexFileInfo 对象列表的字典
        futures = {executor.submit(MergeFiles.merge_files, rinex_files, merge_path): i for i, rinex_files in enumerate(file_groups)}
        # 遍历 future 对象，直到它们完成
        for future in concurrent.futures.as_completed(futures):
            try:
                # 尝试获取 future 的结果，如果出现异常则捕获并打印
                future.result()
            except Exception as e:
                print(f"发生异常：{e}")


directory_path = r"D:\Ropeway\GNSS\FTP\B021"
merge_path = r"D:\Ropeway\GNSS\FTP\Merge"

rnx_files: dict[datetime.date, list[RinexFileInfo]] = get_rnx_files_dict_date(directory_path)

to_process_rnx_file: list[list[RinexFileInfo]] = []

for start_date, rinex_list in rnx_files.items():
    o_files: list[RinexFileInfo] = [rnx_file for rnx_file in rinex_list if rnx_file.file_type == RinexFileType.O]
    n_files: list[RinexFileInfo] = [rnx_file for rnx_file in rinex_list if rnx_file.file_type == RinexFileType.N]
    to_process_rnx_file.append(o_files)
    to_process_rnx_file.append(n_files)


merge_files_threadpool_version2(to_process_rnx_file, merge_path)